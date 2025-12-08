#!/usr/bin/env python3
"""
Direct Scraper to Excel - Extract ALL data from DOM (not __NEXT_DATA__).
Optimized for fast internet + limited server resources.
"""

import csv
import json
import os
import sys
import asyncio
import logging
import re
from datetime import datetime
from playwright.async_api import async_playwright
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S',
    handlers=[logging.StreamHandler(sys.stdout)]
)
log = logging.getLogger(__name__)

# Configuration
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MAX_CONCURRENT = 10  # Can increase since we're waiting for DOM anyway
PAGE_TIMEOUT = 30000  # 30s for full DOM render
ROW_HEIGHT = 80
USE_IMAGE_FORMULA = True

log.info(f"Script directory: {SCRIPT_DIR}")
log.info(f"Config: MAX_CONCURRENT={MAX_CONCURRENT}, PAGE_TIMEOUT={PAGE_TIMEOUT}ms")


# =============================================================================
# DOM EXTRACTION FUNCTIONS - All data from rendered HTML
# =============================================================================

def extract_product_name(soup):
    """Extract product name from DOM"""
    # Try multiple selectors
    selectors = [
        ('h1', {'class': re.compile(r'product.*name', re.I)}),
        ('h1', {}),
        ('span', {'class': re.compile(r'product.*title', re.I)}),
    ]
    for tag, attrs in selectors:
        el = soup.find(tag, attrs)
        if el:
            return el.get_text(strip=True)
    return ''


def extract_product_sku(soup):
    """Extract SKU from DOM"""
    # Look for SKU in various places
    sku_el = soup.find(string=re.compile(r'SKU|Style', re.I))
    if sku_el:
        parent = sku_el.parent
        if parent:
            text = parent.get_text(strip=True)
            match = re.search(r'[A-Z]{2,}\d+', text)
            if match:
                return match.group()
    
    # Try from input fields
    inputs = soup.find_all('input', {'name': True})
    for inp in inputs:
        name = inp.get('name', '')
        if re.match(r'^[A-Z]{2,}\d+', name):
            return name.split('-')[0] if '-' in name else name
    
    return ''


def extract_prices(soup):
    """Extract retail and wholesale prices from DOM"""
    retail = ''
    wholesale = ''
    
    # Look for price elements
    price_container = soup.find('div', class_=re.compile(r'price', re.I))
    if price_container:
        prices = price_container.find_all(string=re.compile(r'\$[\d,]+\.?\d*'))
        if len(prices) >= 2:
            retail = prices[0].strip()
            wholesale = prices[1].strip()
        elif len(prices) == 1:
            retail = prices[0].strip()
    
    # Alternative: look for specific price labels
    if not retail:
        retail_el = soup.find(string=re.compile(r'retail', re.I))
        if retail_el and retail_el.parent:
            price_match = re.search(r'\$[\d,]+\.?\d*', retail_el.parent.get_text())
            if price_match:
                retail = price_match.group()
    
    if not wholesale:
        wholesale_el = soup.find(string=re.compile(r'wholesale', re.I))
        if wholesale_el and wholesale_el.parent:
            price_match = re.search(r'\$[\d,]+\.?\d*', wholesale_el.parent.get_text())
            if price_match:
                wholesale = price_match.group()
    
    return retail, wholesale


def extract_description(soup):
    """Extract product description from DOM"""
    # Try common description containers
    selectors = [
        ('div', {'class': re.compile(r'description|design.?intent', re.I)}),
        ('p', {'class': re.compile(r'description', re.I)}),
        ('div', {'data-testid': re.compile(r'description', re.I)}),
    ]
    for tag, attrs in selectors:
        el = soup.find(tag, attrs)
        if el:
            return el.get_text(strip=True)[:500]  # Limit length
    return ''


def extract_product_type(soup):
    """Extract product type/category from DOM"""
    # Look for breadcrumbs or category labels
    breadcrumb = soup.find('nav', {'aria-label': re.compile(r'breadcrumb', re.I)})
    if breadcrumb:
        items = breadcrumb.find_all(['a', 'span'])
        if len(items) >= 2:
            return items[-2].get_text(strip=True)
    
    # Try category label
    cat_el = soup.find(string=re.compile(r'category|type', re.I))
    if cat_el and cat_el.parent:
        return cat_el.parent.get_text(strip=True).replace('Category:', '').replace('Type:', '').strip()
    
    return ''


def extract_product_image(soup):
    """Extract main product image URL from DOM"""
    # Try primary image selectors
    selectors = [
        ('img', {'class': re.compile(r'image_image|product.*image|main.*image', re.I)}),
        ('img', {'alt': re.compile(r'product', re.I)}),
    ]
    
    for tag, attrs in selectors:
        img = soup.find(tag, attrs)
        if img:
            # Try srcset first for highest quality
            srcset = img.get('srcset', '')
            if srcset:
                # Get largest image from srcset
                best_url = ''
                best_size = 0
                for source in srcset.split(','):
                    parts = source.strip().split(' ')
                    if len(parts) >= 2 and parts[0].startswith('http'):
                        size_match = re.search(r'(\d+)w', parts[1])
                        if size_match:
                            size = int(size_match.group(1))
                            if size > best_size:
                                best_size = size
                                best_url = parts[0]
                if best_url:
                    return best_url
            
            # Fallback to src
            src = img.get('src', '')
            if src.startswith('http'):
                return src
    
    return ''


def extract_color_swatches(soup):
    """Extract color swatches from DOM"""
    swatches = []
    
    # Find swatch container
    container = soup.find('div', class_=re.compile(r'color.*swatch|swatch.*container', re.I))
    if not container:
        container = soup
    
    # Find swatch images
    for img in container.find_all('img', class_=re.compile(r'swatch', re.I)):
        name = img.get('alt', 'Unknown')
        url = ''
        
        # Try srcset first
        srcset = img.get('srcset', '')
        if srcset:
            for src in srcset.split(','):
                parts = src.strip().split(' ')
                if parts and parts[0].startswith('http'):
                    url = parts[0]
                    break
        
        # Fallback to src
        if not url:
            url = img.get('src', '')
        
        if url.startswith('http'):
            swatches.append({'name': name, 'url': url})
    
    return swatches


def extract_inventory(soup):
    """Extract inventory data from DOM - sizes, quantities, SKUs"""
    inventory = {}
    
    # Find inventory accordion/grid
    accordions = soup.find_all('details', class_=re.compile(r'accordion|inventory', re.I))
    
    if not accordions:
        # Try alternative structure
        accordions = soup.find_all('div', class_=re.compile(r'inventory.*grid|color.*section', re.I))
    
    for accordion in accordions:
        # Get color name from heading
        heading = accordion.find(['span', 'div', 'h3', 'h4'], class_=re.compile(r'heading|title|color.*name', re.I))
        if not heading:
            heading = accordion.find('summary')
        
        if not heading:
            continue
        
        color = heading.get_text(strip=True)
        if not color:
            continue
        
        # Find inventory table
        table = accordion.find('table')
        if not table:
            continue
        
        items = []
        tbody = table.find('tbody') or table
        
        for row in tbody.find_all('tr'):
            # Get size
            size_el = row.find(['span', 'td'], class_=re.compile(r'size', re.I))
            if not size_el:
                cells = row.find_all('td')
                if cells:
                    size_el = cells[0]
            
            if not size_el:
                continue
            
            size = size_el.get_text(strip=True)
            if not size:
                continue
            
            # Get quantity
            qty = 0
            qty_els = row.find_all(['span', 'td'], class_=re.compile(r'quantity|qty|stock', re.I))
            for q in qty_els:
                try:
                    qty += int(re.sub(r'[^\d]', '', q.get_text(strip=True)) or 0)
                except:
                    pass
            
            # If no quantity found, try all cells
            if qty == 0:
                for cell in row.find_all('td'):
                    text = cell.get_text(strip=True)
                    if text.isdigit():
                        qty += int(text)
            
            # Get SKU from input
            sku = None
            inp = row.find('input', {'name': True})
            if inp:
                sku = inp.get('name')
            
            items.append({
                'size': size,
                'sku': sku,
                'available_quantity': qty,
                'in_stock': qty > 0
            })
        
        if items:
            inventory[color] = items
    
    return inventory


def extract_slug_from_url(url):
    """Extract product slug from URL"""
    parts = url.rstrip('/').split('/')
    return parts[-1] if parts else ''


def extract_all_from_dom(html, url):
    """Extract ALL product data from DOM only"""
    soup = BeautifulSoup(html, 'html.parser')
    
    name = extract_product_name(soup)
    sku = extract_product_sku(soup)
    retail, wholesale = extract_prices(soup)
    description = extract_description(soup)
    product_type = extract_product_type(soup)
    image = extract_product_image(soup)
    swatches = extract_color_swatches(soup)
    inventory = extract_inventory(soup)
    slug = extract_slug_from_url(url)
    
    # Build product dict similar to what we had before
    product = {
        'name': name,
        'sku': sku,
        'retail': retail,
        'wholesale': wholesale,
        'description': description,
        'product_type': product_type,
        'slug': slug,
    }
    
    # Check if we got meaningful data
    if not name and not inventory:
        return None, None, None, None
    
    return product, inventory, swatches, image


# =============================================================================
# EXCEL FORMATTING FUNCTIONS
# =============================================================================

def setup_worksheet(ws, name):
    """Setup worksheet with headers and formatting"""
    ws.title = name
    widths = {
        'A': 20, 'B': 30, 'C': 20, 'D': 25, 'E': 35, 
        'F': 15, 'G': 15, 'H': 20, 'I': 12, 'J': 12, 
        'K': 20, 'L': 12, 'M': 40, 'N': 25, 'O': 20
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    
    headers = [
        "Product Image", "Product Name", "SKU", "colorSku", "sku name c",
        "Retail Price", "Wholesale Price", "color names", "Current Color Swatch",
        "Size", "quantity", "instock", "Description", "Slug", "product type"
    ]
    
    fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='medium'), right=Side(style='medium'),
        top=Side(style='medium'), bottom=Side(style='medium')
    )
    
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    ws.row_dimensions[1].height = 30


def add_row(ws, row, data):
    """Add a data row to worksheet"""
    thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    fill = PatternFill(
        start_color="E8F4F8" if row % 2 == 0 else "FFFFFF",
        end_color="E8F4F8" if row % 2 == 0 else "FFFFFF",
        fill_type="solid"
    )
    
    # Column A - Product Image
    if data.get('image') and USE_IMAGE_FORMULA:
        ws[f'A{row}'] = f'=IMAGE("{data["image"]}",1)'
        ws.row_dimensions[row].height = ROW_HEIGHT
    
    # Column B-O - Data
    ws[f'B{row}'] = data.get('name', '')
    ws[f'C{row}'] = data.get('sku', '')
    ws[f'D{row}'] = data.get('color_sku', '')
    ws[f'E{row}'] = data.get('sku_name', '')
    ws[f'F{row}'] = data.get('retail', '')
    ws[f'G{row}'] = data.get('wholesale', '')
    ws[f'H{row}'] = data.get('color', '')
    
    # Column I - Color Swatch
    if data.get('swatch_url') and USE_IMAGE_FORMULA:
        ws[f'I{row}'] = f'=IMAGE("{data["swatch_url"]}",1)'
    
    ws[f'J{row}'] = data.get('size', '')
    ws[f'K{row}'] = str(data.get('qty', 0))
    ws[f'L{row}'] = 'Yes' if data.get('in_stock') else 'No'
    ws[f'M{row}'] = data.get('description', '')
    ws[f'N{row}'] = data.get('slug', '')
    ws[f'O{row}'] = data.get('product_type', '')
    
    # Apply formatting
    for cell in ws[row]:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin
        cell.fill = fill


def add_product(ws, row, product, inventory, swatches, image):
    """Add all rows for a product"""
    base = {
        'name': product.get('name', ''),
        'sku': product.get('sku', ''),
        'sku_name': '',
        'retail': product.get('retail', ''),
        'wholesale': product.get('wholesale', ''),
        'description': product.get('description', ''),
        'slug': product.get('slug', ''),
        'product_type': product.get('product_type', ''),
        'image': image
    }
    
    if inventory:
        for color, items in inventory.items():
            # Find matching swatch
            swatch_url = ''
            for s in swatches:
                if s['name'].lower() == color.lower():
                    swatch_url = s['url']
                    break
            
            for item in items:
                data = {
                    **base,
                    'color': color,
                    'swatch_url': swatch_url,
                    'size': item['size'],
                    'qty': item['available_quantity'],
                    'in_stock': item['in_stock']
                }
                
                if item.get('sku'):
                    data['sku'] = item['sku']
                    data['color_sku'] = f"{item['sku']}-{color.lower().replace(' ', '')}"
                
                add_row(ws, row, data)
                row += 1
    else:
        # No inventory - add single row
        add_row(ws, row, {**base, 'color': '', 'size': '', 'qty': 0, 'in_stock': False})
        row += 1
    
    return row


# =============================================================================
# SCRAPER CLASS
# =============================================================================

class Scraper:
    def __init__(self, cookies_file, categories_folder, output_folder):
        self.cookies_file = cookies_file
        self.categories_folder = categories_folder
        self.output_folder = output_folder
        self.cookies = []
        self.data = {}
        self.stats = {'done': 0, 'fail': 0, 'total': 0}
        
        log.info(f"Scraper initialized")
        log.info(f"  Cookies: {cookies_file}")
        log.info(f"  Categories: {categories_folder}")
        log.info(f"  Output: {output_folder}")
    
    def check_prerequisites(self):
        """Check required files exist"""
        log.info("Checking prerequisites...")
        errors = []
        
        if not os.path.exists(self.cookies_file):
            errors.append(f"Cookie file NOT FOUND: {self.cookies_file}")
            log.error(f"❌ Cookie file missing: {self.cookies_file}")
            log.error(f"   Run: python login_and_save_cookies.py")
        else:
            log.info(f"✓ Cookie file exists")
        
        if not os.path.exists(self.categories_folder):
            errors.append(f"Categories folder NOT FOUND: {self.categories_folder}")
            log.error(f"❌ Categories folder missing: {self.categories_folder}")
        else:
            csv_files = [f for f in os.listdir(self.categories_folder) if f.endswith('.csv')]
            if csv_files:
                log.info(f"✓ Categories folder: {csv_files}")
            else:
                errors.append("No CSV files in categories folder")
                log.error(f"❌ No CSV files found")
        
        if errors:
            log.error("="*50)
            log.error("PREREQUISITES FAILED - Run these first:")
            log.error("  1. python login_and_save_cookies.py")
            log.error("  2. python extract_product_links.py")
            log.error("="*50)
            return False
        
        return True
    
    def load_cookies(self):
        """Load cookies from file"""
        log.info(f"Loading cookies...")
        with open(self.cookies_file) as f:
            for c in json.load(f):
                cookie = {
                    'name': c['name'],
                    'value': c['value'],
                    'domain': c.get('domain', '.lululemon.com'),
                    'path': c.get('path', '/')
                }
                if c.get('sameSite') in ['Strict', 'Lax', 'None']:
                    cookie['sameSite'] = c['sameSite']
                self.cookies.append(cookie)
        log.info(f"✓ Loaded {len(self.cookies)} cookies")
    
    def load_urls(self, cat):
        """Load product URLs for a category"""
        path = os.path.join(self.categories_folder, f"{cat}.csv")
        if not os.path.exists(path):
            return []
        with open(path) as f:
            reader = csv.reader(f)
            next(reader, None)  # Skip header
            return [r[0] for r in reader if r]
    
    async def scrape_page(self, context, url, cat, sem):
        """Scrape a single product page - DOM only"""
        pid = url.rstrip('/').split('/')[-1]
        
        async with sem:
            page = None
            try:
                page = await context.new_page()
                
                # Block heavy resources for speed
                await page.route('**/*', lambda r: (
                    r.abort() if r.request.resource_type in ['image', 'media', 'font'] 
                    or any(p in r.request.url.lower() for p in [
                        'analytics', 'facebook', 'google-analytics', 'gtm',
                        'doubleclick', 'hotjar', '.woff', '.woff2'
                    ])
                    else r.continue_()
                ))
                
                # Load page and wait for DOM to be ready
                log.debug(f"[{cat}] Loading: {pid}")
                await page.goto(url, wait_until='domcontentloaded', timeout=PAGE_TIMEOUT)
                
                # Wait for inventory table to appear (key element)
                try:
                    await page.wait_for_selector('table', timeout=10000)
                    log.debug(f"[{cat}] Table found: {pid}")
                except:
                    log.debug(f"[{cat}] No table, continuing: {pid}")
                
                # Small delay to ensure DOM is stable
                await asyncio.sleep(0.5)
                
                # Get rendered HTML
                html = await page.content()
                await page.close()
                
                # Extract ALL data from DOM
                product, inventory, swatches, image = extract_all_from_dom(html, url)
                
                if product:
                    self.data[cat].append((product, inventory, swatches, image))
                    self.stats['done'] += 1
                    
                    inv_count = sum(len(items) for items in inventory.values()) if inventory else 0
                    log.debug(f"[{cat}] ✓ {pid}: {product.get('name', '')[:30]}... ({inv_count} inventory items)")
                    
                    if self.stats['done'] % 20 == 0:
                        log.info(f"Progress: {self.stats['done']}/{self.stats['total']} done, {self.stats['fail']} failed")
                    
                    return True
                
                self.stats['fail'] += 1
                log.warning(f"[{cat}] ✗ No data extracted: {pid}")
                return False
                
            except Exception as e:
                self.stats['fail'] += 1
                log.error(f"[{cat}] ✗ {pid}: {type(e).__name__}: {str(e)[:50]}")
                if page:
                    try:
                        await page.close()
                    except:
                        pass
                return False
    
    async def run(self):
        """Main scraping loop"""
        log.info("="*50)
        log.info("SCRAPE TO EXCEL - DOM EXTRACTION")
        log.info("All data extracted from rendered DOM")
        log.info("="*50)
        
        if not self.check_prerequisites():
            return
        
        self.load_cookies()
        
        # Load all category URLs
        cats = ['women', 'men', 'accessories', 'supplies']
        urls = {}
        
        log.info("\nLoading categories...")
        for c in cats:
            urls[c] = self.load_urls(c)
            self.stats['total'] += len(urls[c])
            if urls[c]:
                log.info(f"  {c}: {len(urls[c])} products")
        
        log.info(f"\nTotal: {self.stats['total']} products to scrape")
        
        if self.stats['total'] == 0:
            log.error("No products found! Check category CSV files.")
            return
        
        start = datetime.now()
        
        try:
            async with async_playwright() as p:
                log.info("\nLaunching browser...")
                browser = await p.chromium.launch(
                    headless=True,
                    args=[
                        '--headless=new',
                        '--disable-gpu',
                        '--no-sandbox',
                        '--disable-dev-shm-usage',
                        '--disable-web-security'
                    ]
                )
                log.info(f"✓ Browser ready")
                log.info(f"  Concurrency: {MAX_CONCURRENT}")
                log.info(f"  Timeout: {PAGE_TIMEOUT}ms")
                
                for cat in cats:
                    if not urls[cat]:
                        continue
                    
                    log.info(f"\n{'='*40}")
                    log.info(f"CATEGORY: {cat.upper()} ({len(urls[cat])} products)")
                    log.info(f"{'='*40}")
                    
                    self.data[cat] = []
                    
                    ctx = await browser.new_context(java_script_enabled=True)
                    await ctx.add_cookies(self.cookies)
                    
                    sem = asyncio.Semaphore(MAX_CONCURRENT)
                    cat_start = datetime.now()
                    
                    await asyncio.gather(*[
                        self.scrape_page(ctx, u, cat, sem) 
                        for u in urls[cat]
                    ])
                    
                    cat_dur = (datetime.now() - cat_start).total_seconds()
                    log.info(f"✓ {cat}: {len(self.data[cat])}/{len(urls[cat])} in {cat_dur:.1f}s")
                    
                    await ctx.close()
                
                await browser.close()
                log.info("\n✓ Browser closed")
        
        except Exception as e:
            log.error(f"Browser error: {type(e).__name__}: {str(e)}")
            raise
        
        # Write Excel
        log.info("\n" + "="*40)
        log.info("WRITING EXCEL FILE")
        log.info("="*40)
        
        wb = Workbook()
        first = True
        total_rows = 0
        
        for cat in cats:
            if not self.data.get(cat):
                continue
            
            ws = wb.active if first else wb.create_sheet()
            first = False
            setup_worksheet(ws, cat.capitalize())
            
            row = 2
            for product, inventory, swatches, image in self.data[cat]:
                row = add_product(ws, row, product, inventory, swatches, image)
            
            rows_in_cat = row - 2
            total_rows += rows_in_cat
            log.info(f"  {cat}: {len(self.data[cat])} products → {rows_in_cat} rows")
        
        # Save file
        os.makedirs(self.output_folder, exist_ok=True)
        out_file = os.path.join(
            self.output_folder, 
            f"lululemon_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        wb.save(out_file)
        
        # Summary
        dur = (datetime.now() - start).total_seconds()
        log.info("\n" + "="*50)
        log.info("COMPLETE!")
        log.info("="*50)
        log.info(f"  Products scraped: {self.stats['done']}/{self.stats['total']}")
        log.info(f"  Failed: {self.stats['fail']}")
        log.info(f"  Total Excel rows: {total_rows}")
        log.info(f"  Time: {dur:.0f}s")
        log.info(f"  Speed: {self.stats['done']/max(dur,1):.1f} products/sec")
        log.info(f"  Output: {out_file}")
        log.info("="*50)


async def main():
    log.info("Starting scrape_to_excel.py (DOM extraction)")
    log.info(f"Working directory: {os.getcwd()}")
    
    scraper = Scraper(
        cookies_file=os.path.join(SCRIPT_DIR, "data/cookie/cookie.json"),
        categories_folder=os.path.join(SCRIPT_DIR, "data/categories"),
        output_folder=os.path.join(SCRIPT_DIR, "data/results")
    )
    
    await scraper.run()


if __name__ == "__main__":
    asyncio.run(main())
