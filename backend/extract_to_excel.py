#!/usr/bin/env python3
"""
Product Information Extractor to Excel
Extracts and formats product information from Lululemon wholesale HTML to Excel with image URLs
Processes all HTML files in the 'web' folder and generates a single Excel sheet

NOTE: The =IMAGE() formula works in:
  - Google Sheets (all versions) ✓
  - Excel 365 / Excel 2021+     # Column S: Color Names (all colors as comma-separated text)
    if color_swatches:
        all_color_names = ", ".join([swatch['name'] for swatch in color_swatches])
        ws[f'S{row_num}'] = all_color_names
        ws[f'S{row_num}'].alignment = Alignment(wrap_text=True, vertical='center')
    else:
        ws[f'S{row_num}'] = color_names
    
    # Column T: Gender
    ws[f'T{row_num}'] = gender
    
    # Column U: Product Type
    ws[f'U{row_num}'] = product_typexcel versions will show the formula text instead of images
  
RECOMMENDATION: Upload the generated .xlsx file to Google Sheets for best compatibility
"""

import json
import os
import glob
from datetime import datetime
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ============================================================
# CONFIGURATION
# ============================================================
# Use relative paths for portability (works on local machine and server)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
WEB_FOLDER = os.path.join(SCRIPT_DIR, "web")  # Folder containing HTML files
DATA_FOLDER = os.path.join(SCRIPT_DIR, "data", "html")  # HTML files from download_by_category.py
ROW_HEIGHT = 80             # Row height in Excel units (points) - reduced for compact view

# Set to True to use IMAGE formulas (works in Google Sheets & Excel 365+)
# Set to False to use clickable hyperlinks (works in all Excel versions)
USE_IMAGE_FORMULA = True

# IMAGE SIZE CONFIGURATION
PRODUCT_IMAGE_WIDTH = 150   # Width in pixels for main product image
PRODUCT_IMAGE_HEIGHT = 150  # Height in pixels for main product image
COLOR_SWATCH_WIDTH = 40     # Width in pixels for color swatch images (smaller, more compact)
COLOR_SWATCH_HEIGHT = 40    # Height in pixels for color swatch images (smaller, more compact)
IMAGE_SIZING_MODE = 0       # Use 0 for best results

# Maximum number of color swatch columns to create
MAX_COLOR_SWATCHES = 6      # Will create columns F through K for color swatches

# ============================================================


def extract_product_image(soup):
    """Extract main product image from the HTML"""
    # Find the main product image with class 'image_image__ECDWj'
    img_tag = soup.find('img', class_='image_image__ECDWj')
    
    if img_tag:
        # Try to get from srcset (prefer higher quality)
        srcset = img_tag.get('srcset', '')
        if srcset:
            # Parse srcset and get a medium-sized image (1280w or 1080w)
            for source in srcset.split(','):
                parts = source.strip().split(' ')
                if len(parts) >= 2 and parts[0].startswith('http'):
                    # Look for 1280w or 1080w size
                    if '1280w' in parts[1] or '1080w' in parts[1]:
                        return parts[0]
            
            # If specific size not found, return first valid URL
            for source in srcset.split(','):
                parts = source.strip().split(' ')
                if parts and parts[0].startswith('http'):
                    return parts[0]
        
        # Fallback to src attribute
        src = img_tag.get('src', '')
        if src and src.startswith('http'):
            return src
    
    return ''


def extract_color_swatches(soup):
    """Extract color swatch images from the HTML"""
    color_swatches = []
    
    # Find the color swatches container
    swatch_container = soup.find('div', class_='color-swatches-selector_colorSwatchContainer__fjw54')
    
    if swatch_container:
        # Find all color swatch images
        swatch_images = swatch_container.find_all('img', class_='color-swatch_colorSwatchImg__apmdW')
        
        for img in swatch_images:
            color_name = img.get('alt', 'Unknown')
            
            # Get the src attribute directly (it should be the full URL)
            img_url = img.get('src')
            
            # If src is not available, try to parse srcset
            if not img_url or not img_url.startswith('http'):
                srcset = img.get('srcset', '')
                if srcset:
                    # Parse srcset and get the first valid URL
                    for source in srcset.split(','):
                        parts = source.strip().split(' ')
                        if parts and parts[0].startswith('http'):
                            img_url = parts[0]
                            break
            
            if img_url and img_url.startswith('http'):
                color_swatches.append({
                    'name': color_name,
                    'url': img_url
                })
    
    return color_swatches


def extract_inventory_from_html(soup):
    """Extract inventory/stock information organized by color, summing quantities across all lots"""
    
    inventory_by_color = {}
    
    # Find all inventory accordion items (one per color)
    accordion_items = soup.find_all('details', class_='inventory-grid_accordionItem__XXIck')
    
    for accordion in accordion_items:
        # Get color name from accordion header
        color_heading = accordion.find('span', class_='inventory-grid_accordionHeadingContent__oebUk')
        if not color_heading:
            continue
        
        color_name = color_heading.get_text(strip=True)
        
        # Find the table for this color
        table = accordion.find('table')
        if not table:
            continue
        
        # Extract inventory data for each size, summing across all lots
        color_inventory = []
        rows = table.find('tbody').find_all('tr') if table.find('tbody') else []
        
        for row in rows:
            # Find size
            size_span = row.find('span', class_='inventory-grid-table_size__5wMgv')
            if not size_span:
                continue
            
            size = size_span.get_text(strip=True)
            
            # Find ALL quantity spans in this row (one per lot)
            quantity_spans = row.find_all('span', class_='inventory-grid-table_quantity__Q0EiU')
            
            # Find ALL input fields (one per lot) for SKU - we'll use the first one
            input_fields = row.find_all('input', {'name': True})
            
            # Sum quantities across all lots for this size
            total_quantity = 0
            for quantity_span in quantity_spans:
                quantity_text = quantity_span.get_text(strip=True)
                try:
                    quantity = int(quantity_text)
                    total_quantity += quantity
                except (ValueError, TypeError):
                    pass
            
            # Get the first SKU (they're all for the same size, just different lots)
            sku = input_fields[0].get('name') if input_fields else None
            
            # Create a single entry for this size with summed quantity
            color_inventory.append({
                "size": size,
                "sku": sku,
                "available_quantity": total_quantity,
                "in_stock": total_quantity > 0
            })
        
        inventory_by_color[color_name] = color_inventory
    
    return inventory_by_color


def extract_product_from_html(html_file_path):
    """Extract product data from local HTML file"""
    
    print("=" * 60)
    print("Lululemon Product Information Extractor to Excel")
    print("=" * 60)
    
    # Read HTML file
    with open(html_file_path, 'r', encoding='utf-8') as f:
        html_content = f.read()
    
    print(f"Reading HTML from: {html_file_path}")
    
    # Parse HTML
    soup = BeautifulSoup(html_content, 'html.parser')
    
    # Find __NEXT_DATA__ script tag
    next_data_script = soup.find('script', {'id': '__NEXT_DATA__', 'type': 'application/json'})
    
    if not next_data_script:
        print("❌ Could not find __NEXT_DATA__ script tag")
        return None, None, None
    
    # Parse JSON
    try:
        data = json.loads(next_data_script.string)
        print("✓ Successfully parsed __NEXT_DATA__ JSON")
    except json.JSONDecodeError as e:
        print(f"❌ Error parsing JSON: {e}")
        return None, None, None
    
    # Navigate to product data
    try:
        product_data = data['props']['pageProps']['data']['pageFolder']['dataSourceConfigurations'][0]['preloadedValue']['product']
        print("✓ Found product data")
    except (KeyError, IndexError, TypeError) as e:
        print(f"❌ Error navigating to product data: {e}")
        return None, None, None, None
    
    # Extract inventory data from HTML
    inventory_data = extract_inventory_from_html(soup)
    if inventory_data:
        print(f"✓ Found inventory data for {len(inventory_data)} size(s)")
    else:
        print("⚠ No inventory data found in HTML")
    
    # Extract product image from HTML
    product_image_url = extract_product_image(soup)
    if product_image_url:
        print(f"✓ Found product image")
    else:
        print("⚠ No product image found")
    
    # Extract color swatches
    color_swatches = extract_color_swatches(soup)
    if color_swatches:
        print(f"✓ Found {len(color_swatches)} color swatch(es)")
    else:
        print("⚠ No color swatches found")
    
    return product_data, inventory_data, color_swatches, product_image_url


def add_product_rows(ws, row_num, product, inventory_data, color_swatches, product_image_url):
    """Add multiple product rows to the worksheet - one row per color/size/lot combination"""
    
    # Extract product data
    variants = product.get('variants', [])
    variant = variants[0] if variants else {}
    attributes = variant.get('attributes', {})
    
    # Get all the required data (common to all rows)
    product_name = product.get('name', '')
    sku = variant.get('sku', '')
    retail_price = f"${product.get('retailPriceRange', ['0'])[0]}"
    wholesale_price = f"${product.get('wholesalePriceRange', ['0'])[0]}"
    description = variant.get('designIntent', '')
    slug = product.get('slug', '')
    sku_name = attributes.get('skuName', '')
    color_code = attributes.get('colourCode', '')
    color_name = attributes.get('colourName', '')
    color_description = attributes.get('colourDescription', '')
    gender = ', '.join(attributes.get('gender', [])) if isinstance(attributes.get('gender'), list) else attributes.get('gender', '')
    product_type = ', '.join(attributes.get('productType', [])) if isinstance(attributes.get('productType'), list) else attributes.get('productType', '')
    size = attributes.get('size', '')
    active = attributes.get('active', False)
    
    # Combine color information
    color_names = f"{color_name} - {color_description}" if color_description else color_name
    
    # If we have inventory data, create a row for each color/size/lot combination
    if inventory_data:
        # Calculate total quantity across all colors and sizes for this product
        total_product_quantity = 0
        for color, inventory_items in inventory_data.items():
            for inv_item in inventory_items:
                total_product_quantity += inv_item.get('available_quantity', 0)
        
        rows_added = 0
        for color, inventory_items in inventory_data.items():
            for inv_item in inventory_items:
                is_last_row = (rows_added == sum(len(items) for items in inventory_data.values()) - 1)
                # Add a row for this specific size/color/lot combination
                add_single_inventory_row(
                    ws, row_num + rows_added, product_name, sku, retail_price, wholesale_price,
                    color_swatches, product_image_url, description, slug, sku_name,
                    color_names, product_type, color, inv_item,
                    total_product_quantity, is_last_row
                )
                rows_added += 1
        
        return row_num + rows_added  # Return next row number
    else:
        # If no inventory data, add a single row with basic product info
        add_single_inventory_row(
            ws, row_num, product_name, sku, retail_price, wholesale_price,
            color_swatches, product_image_url, description, slug, sku_name,
            color_names, product_type, color_name, None, 0, True
        )
        return row_num + 1


def add_single_inventory_row(ws, row_num, product_name, sku, retail_price, wholesale_price,
                              color_swatches, product_image_url, description, slug, sku_name,
                              color_names, product_type, current_color, inv_item,
                              total_product_quantity, is_last_row):
    """Add a single row with specific inventory details"""
    
    # Define border styles
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Prominent border for last row of each product (darker and thicker)
    if is_last_row:
        bottom_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='medium', color='2E4053')  # Dark blue-gray, medium thickness
        )
    else:
        bottom_border = thin_border
    
    # Improved alternating row colors for better readability
    if row_num % 2 == 0:
        fill_color = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")  # Light blue
    else:
        fill_color = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White
    
    # Column A: Product Image (with IMAGE formula - Google Sheets compatible)
    if product_image_url:
        if USE_IMAGE_FORMULA:
            ws[f'A{row_num}'] = f'=IMAGE("{product_image_url}",1)'  # Added mode 1 for fit-to-cell
        else:
            # Use hyperlink for older Excel compatibility
            ws[f'A{row_num}'].hyperlink = product_image_url
            ws[f'A{row_num}'].value = "View Image"
            ws[f'A{row_num}'].style = "Hyperlink"
        ws.row_dimensions[row_num].height = ROW_HEIGHT
    
    # Column B: Product Name
    ws[f'B{row_num}'] = product_name
    
    # Column C: SKU (use the variant-specific SKU if available from inventory)
    if inv_item and inv_item.get('sku'):
        ws[f'C{row_num}'] = inv_item['sku']
    else:
        ws[f'C{row_num}'] = sku
    
    # Column D: colorSku (format: SKU-ColorName, e.g., "144525671-Black")
    if inv_item and inv_item.get('sku') and current_color:
        color_formatted = current_color.lower().replace(' ', '')
        ws[f'D{row_num}'] = f"{inv_item['sku']}-{color_formatted}"
    else:
        ws[f'D{row_num}'] = ''
    
    # Column E: SKU Name c
    ws[f'E{row_num}'] = sku_name
    
    # Column F: Retail Price
    ws[f'F{row_num}'] = retail_price
    
    # Column G: Wholesale Price
    ws[f'G{row_num}'] = wholesale_price
    
    # Column H: color names (only the active color for this row)
    ws[f'H{row_num}'] = current_color if current_color else color_names
    
    # Column I: Current Color Swatch Image (only the specific color for this row)
    if color_swatches and current_color:
        # Find the matching color swatch for the current color
        matching_swatch = None
        for swatch in color_swatches:
            # Case-insensitive comparison
            if swatch['name'].lower() == current_color.lower():
                matching_swatch = swatch
                break
        
        # Display only the matching color swatch
        if matching_swatch:
            if USE_IMAGE_FORMULA:
                ws[f'I{row_num}'] = f'=IMAGE("{matching_swatch["url"]}",1)'  # Added mode 1 for fit-to-cell
            else:
                ws[f'I{row_num}'].hyperlink = matching_swatch["url"]
                ws[f'I{row_num}'].value = matching_swatch["name"][:10]
                ws[f'I{row_num}'].style = "Hyperlink"
    
    # Column J: Size (specific to this row)
    if inv_item:
        ws[f'J{row_num}'] = inv_item.get('size', '')
    else:
        ws[f'J{row_num}'] = ''
    
    # Column K: Available Quantity (format: individual/total, e.g., "50/439")
    if inv_item:
        individual_qty = inv_item.get('available_quantity', 0)
        ws[f'K{row_num}'] = f"{individual_qty}"
    else:
        ws[f'K{row_num}'] = f"0/{total_product_quantity}"
    
    # Column L: In Stock (for this specific size/color)
    if inv_item:
        ws[f'L{row_num}'] = 'Yes' if inv_item.get('in_stock') else 'No'
    else:
        ws[f'L{row_num}'] = 'No'
    
    # Column M: Description
    ws[f'M{row_num}'] = description
    
    # Column N: Slug
    ws[f'N{row_num}'] = slug
    
    # Column O: Product Type (Gender column removed)
    ws[f'O{row_num}'] = product_type
    
    # Apply alignment, borders, and fill color to all cells in the row
    for cell in ws[row_num]:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = bottom_border
        cell.fill = fill_color


def setup_worksheet(ws, sheet_name):
    """Set up a worksheet with headers and column widths"""
    ws.title = sheet_name
    
    # Set column widths (Gender column removed)
    ws.column_dimensions['A'].width = 20  # Product Image (reduced)
    ws.column_dimensions['B'].width = 30  # Product Name
    ws.column_dimensions['C'].width = 20  # SKU
    ws.column_dimensions['D'].width = 25  # colorSku
    ws.column_dimensions['E'].width = 35  # SKU Name c
    ws.column_dimensions['F'].width = 15  # Retail Price
    ws.column_dimensions['G'].width = 15  # Wholesale Price
    ws.column_dimensions['H'].width = 20  # color names (active color only)
    ws.column_dimensions['I'].width = 12  # Current Color Swatch (reduced)
    ws.column_dimensions['J'].width = 12  # Size
    ws.column_dimensions['K'].width = 20  # Available Quantity (format: individual/total)
    ws.column_dimensions['L'].width = 12  # In Stock
    ws.column_dimensions['M'].width = 40  # Description
    ws.column_dimensions['N'].width = 25  # Slug
    ws.column_dimensions['O'].width = 20  # Product Type
    
    # Header style with borders
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Darker professional blue
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_border = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )
    
    # Create headers (Gender column removed)
    headers = [
        "Product Image",        # A
        "Product Name",         # B
        "SKU",                  # C
        "colorSku",             # D (format: SKU-colorname, e.g., "144525671-black")
        "sku name c",           # E
        "Retail Price",         # F
        "Wholesale Price",      # G
        "color names",          # H (only the active color for this row)
        "Current Color Swatch", # I (only the color for this specific row)
        "Size",                 # J (specific size for this row)
        "quantity",     # K (format: individual/total, e.g., "50/439")
        "instock",              # L (in stock status for this specific item)
        "Description",          # M
        "Slug",                 # N
        "product type"          # O (Gender removed)
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = header_border
    
    # Set row height for header
    ws.row_dimensions[1].height = 30


def add_summary_product_row(ws, row_num, product, inventory_data, color_swatches, product_image_url, category):
    """Add a single summary row (one per product) - legacy format"""
    
    # Define border style
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Alternating row colors - using same scheme as detailed sheets
    if row_num % 2 == 0:
        fill_color = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")  # Light blue
    else:
        fill_color = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White
    
    # Extract product data
    variants = product.get('variants', [])
    variant = variants[0] if variants else {}
    attributes = variant.get('attributes', {})
    
    product_name = product.get('name', '')
    sku = variant.get('sku', '')
    retail_price = f"${product.get('retailPriceRange', ['0'])[0]}"
    wholesale_price = f"${product.get('wholesalePriceRange', ['0'])[0]}"
    description = variant.get('designIntent', '')
    slug = product.get('slug', '')
    sku_name = attributes.get('skuName', '')
    color_name = attributes.get('colourName', '')
    color_description = attributes.get('colourDescription', '')
    product_type = ', '.join(attributes.get('productType', [])) if isinstance(attributes.get('productType'), list) else attributes.get('productType', '')
    size = attributes.get('size', '')
    active = attributes.get('active', False)
    
    color_names = f"{color_name} - {color_description}" if color_description else color_name
    
    # Column A: Product Image
    if product_image_url:
        if USE_IMAGE_FORMULA:
            ws[f'A{row_num}'] = f'=IMAGE("{product_image_url}",1)'
        else:
            ws[f'A{row_num}'].hyperlink = product_image_url
            ws[f'A{row_num}'].value = "View Image"
            ws[f'A{row_num}'].style = "Hyperlink"
        ws.row_dimensions[row_num].height = ROW_HEIGHT
    
    # Column B: Product Name
    ws[f'B{row_num}'] = product_name
    
    # Column C: SKU
    ws[f'C{row_num}'] = sku
    
    # Column D: Retail Price
    ws[f'D{row_num}'] = retail_price
    
    # Column E: Wholesale Price
    ws[f'E{row_num}'] = wholesale_price
    
    # Columns F-K: Color Swatches (one image per column, up to 6 colors)
    color_columns = ['F', 'G', 'H', 'I', 'J', 'K']
    if color_swatches:
        for idx, swatch in enumerate(color_swatches[:6]):  # Limit to 6
            col_letter = color_columns[idx]
            if USE_IMAGE_FORMULA:
                ws[f'{col_letter}{row_num}'] = f'=IMAGE("{swatch["url"]}",1)'
            else:
                ws[f'{col_letter}{row_num}'].hyperlink = swatch["url"]
                ws[f'{col_letter}{row_num}'].value = swatch["name"][:10]
                ws[f'{col_letter}{row_num}'].style = "Hyperlink"
    
    # Process inventory data
    if inventory_data:
        all_sizes = []
        total_quantity = 0
        in_stock_items = 0
        
        for color, inventory_items in inventory_data.items():
            for inv_item in inventory_items:
                inv_size = inv_item.get('size', '')
                if inv_size and inv_size not in all_sizes:
                    all_sizes.append(inv_size)
                
                qty = inv_item.get('available_quantity', 0)
                if isinstance(qty, (int, float)):
                    total_quantity += qty
                
                if inv_item.get('in_stock'):
                    in_stock_items += 1
        
        # Column L: All sizes comma-separated
        ws[f'L{row_num}'] = ", ".join(all_sizes) if all_sizes else size
        
        # Column M: Total Available Quantity
        ws[f'M{row_num}'] = total_quantity
        
        # Column P: In Stock
        ws[f'P{row_num}'] = 'Yes' if in_stock_items > 0 else 'No'
    else:
        ws[f'L{row_num}'] = size
        ws[f'M{row_num}'] = 0
        ws[f'P{row_num}'] = 'Yes' if active else 'No'
    
    # Column N: Description
    ws[f'N{row_num}'] = description
    
    # Column O: Slug
    ws[f'O{row_num}'] = slug
    
    # Column Q: SKU Name
    ws[f'Q{row_num}'] = sku_name
    
    # Column R: Color Names (all colors as comma-separated text)
    if color_swatches:
        all_color_names = ", ".join([swatch['name'] for swatch in color_swatches])
        ws[f'R{row_num}'] = all_color_names
    else:
        ws[f'R{row_num}'] = color_names
    
    # Column S: Product Type
    ws[f'S{row_num}'] = product_type
    
    # Column T: Category
    ws[f'T{row_num}'] = category.capitalize()
    
    # Apply alignment, borders, and fill color to all cells in the row
    for cell in ws[row_num]:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        cell.fill = fill_color


def setup_summary_worksheet(ws):
    """Set up the Summary worksheet with headers and column widths"""
    ws.title = "Summary"
    
    # Set column widths
    ws.column_dimensions['A'].width = 20  # Product Image
    ws.column_dimensions['B'].width = 30  # Product Name
    ws.column_dimensions['C'].width = 20  # SKU
    ws.column_dimensions['D'].width = 15  # Retail Price
    ws.column_dimensions['E'].width = 15  # Wholesale Price
    # Columns F-K: Color Swatches
    for col_letter in ['F', 'G', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[col_letter].width = 12
    ws.column_dimensions['L'].width = 12  # Size
    ws.column_dimensions['M'].width = 18  # Total Quantity
    ws.column_dimensions['N'].width = 40  # Description
    ws.column_dimensions['O'].width = 25  # Slug
    ws.column_dimensions['P'].width = 12  # In Stock
    ws.column_dimensions['Q'].width = 35  # SKU Name
    ws.column_dimensions['R'].width = 20  # Color Names
    ws.column_dimensions['S'].width = 20  # Product Type
    ws.column_dimensions['T'].width = 15  # Category
    
    # Header style with borders
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_border = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )
    
    # Create headers
    headers = [
        "Product Image",         # A
        "Product Name",          # B
        "SKU",                   # C
        "Retail Price",          # D
        "Wholesale Price",       # E
        "Color 1",               # F
        "Color 2",               # G
        "Color 3",               # H
        "Color 4",               # I
        "Color 5",               # J
        "Color 6",               # K
        "Sizes",                 # L
        "Total Quantity",        # M
        "Description",           # N
        "Slug",                  # O
        "In Stock",              # P
        "SKU Name",              # Q
        "Color Names",           # R
        "Product Type",          # S
        "Category"               # T
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = header_border
    
    # Set row height for header
    ws.row_dimensions[1].height = 30


def create_excel_with_all_products(all_products_data, output_file):
    """Create Excel file with separate sheets for each category + Summary sheet"""
    
    print("\n" + "=" * 60)
    print("Creating Excel file with category sheets + Summary...")
    print("=" * 60)
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create Summary sheet FIRST (will be first tab)
    summary_ws = wb.create_sheet(title="Summary", index=0)
    setup_summary_worksheet(summary_ws)
    summary_row = 2
    
    # Create sheets for each category
    categories = ['accessories', 'men', 'women', 'supplies']
    sheets = {}
    
    for category in categories:
        ws = wb.create_sheet(title=category.capitalize())
        setup_worksheet(ws, category.capitalize())
        sheets[category] = {'sheet': ws, 'current_row': 2, 'count': 0}
    
    # Add products to their respective category sheets AND summary sheet
    for product_data in all_products_data:
        category = product_data.get('category', 'women')  # default to women if not specified
        
        product = product_data['product']
        inventory = product_data['inventory']
        swatches = product_data['swatches']
        image_url = product_data['image_url']
        
        # Add to category sheet (detailed view)
        if category in sheets:
            ws = sheets[category]['sheet']
            current_row = sheets[category]['current_row']
            new_row = add_product_rows(ws, current_row, product, inventory, swatches, image_url)
            sheets[category]['current_row'] = new_row
            sheets[category]['count'] += 1
        
        # Add to Summary sheet (one row per product)
        add_summary_product_row(summary_ws, summary_row, product, inventory, swatches, image_url, category)
        summary_row += 1
    
    # Print summary
    print("\nProducts added to sheets:")
    print(f"  Summary: {len(all_products_data)} product(s)")
    for category in categories:
        count = sheets[category]['count']
        print(f"  {category.capitalize()}: {count} product(s)")
    
    # Save workbook
    wb.save(output_file)
    print(f"\n✓ Excel file saved to: {output_file}")


def main():
    """Process all HTML files in the web folder and create a single Excel file"""
    
    print("=" * 60)
    print("Lululemon Product Information Batch Extractor")
    print("=" * 60)
    
    # Check if web folder exists, if not try data folder
    html_files = []
    
    if os.path.exists(WEB_FOLDER):
        html_files = glob.glob(os.path.join(WEB_FOLDER, "*.html"))
        print(f"\n✓ Checking web folder: {WEB_FOLDER}")
    
    if not html_files and os.path.exists(DATA_FOLDER):
        print(f"\n✓ Checking data folder: {DATA_FOLDER}")
        # Look in all subfolders of data folder and track category (dynamically discover)
        html_files_with_category = []
        
        # Discover all subfolders in data/html/
        if os.path.isdir(DATA_FOLDER):
            subfolders = [f for f in os.listdir(DATA_FOLDER) 
                         if os.path.isdir(os.path.join(DATA_FOLDER, f))]
            
            for subfolder in subfolders:
                subfolder_path = os.path.join(DATA_FOLDER, subfolder)
                subfolder_files = glob.glob(os.path.join(subfolder_path, "*.html"))
                # Store tuple of (file_path, category)
                for file_path in subfolder_files:
                    html_files_with_category.append((file_path, subfolder))
                if subfolder_files:
                    print(f"  Found {len(subfolder_files)} files in {subfolder}")
        
        # Convert to simple list for counting, keep category mapping
        html_files = [f[0] for f in html_files_with_category]
        # Create a mapping of file to category
        file_category_map = {f[0]: f[1] for f in html_files_with_category}
    else:
        file_category_map = {}
    
    if not html_files:
        print(f"\n❌ No HTML files found in {WEB_FOLDER} or {DATA_FOLDER}")
        print(f"Creating web folder: {WEB_FOLDER}")
        os.makedirs(WEB_FOLDER, exist_ok=True)
        print(f"✓ Folder created. Please add HTML files to: {WEB_FOLDER}")
        return
    
    print(f"\n✓ Found {len(html_files)} HTML file(s) to process")
    print("-" * 60)
    
    # Process all HTML files
    all_products_data = []
    successful = 0
    failed = 0
    
    for html_file in html_files:
        filename = os.path.basename(html_file)
        print(f"\nProcessing: {filename}")
        
        # Determine category from file path
        category = file_category_map.get(html_file, 'women')  # default to women
        
        try:
            product, inventory_data, color_swatches, product_image_url = extract_product_from_html(html_file)
            
            if product:
                all_products_data.append({
                    'product': product,
                    'inventory': inventory_data,
                    'swatches': color_swatches,
                    'image_url': product_image_url,
                    'category': category
                })
                successful += 1
                print(f"  ✓ Successfully extracted: {product.get('name', 'Unknown')} [Category: {category}]")
            else:
                failed += 1
                print(f"  ❌ Failed to extract product data")
        except Exception as e:
            failed += 1
            print(f"  ❌ Error processing file: {e}")
    
    # Check if we have any products to export
    if not all_products_data:
        print("\n❌ No products were successfully extracted")
        return
    
    # Create results directory if it doesn't exist
    results_folder = os.path.join(SCRIPT_DIR, "data", "results")
    os.makedirs(results_folder, exist_ok=True)
    
    # Generate output filename with timestamp (save in data/results/)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(results_folder, f"all_products_{timestamp}.xlsx")
    
    # Create Excel file with all products
    create_excel_with_all_products(all_products_data, output_file)
    
    # Print summary
    print("\n" + "=" * 60)
    print("BATCH EXTRACTION COMPLETE")
    print("=" * 60)
    print(f"Total HTML files processed: {len(html_files)}")
    print(f"Successfully extracted: {successful}")
    print(f"Failed: {failed}")
    print(f"Total products in Excel: {len(all_products_data)}")
    print("=" * 60)
    
    return output_file


if __name__ == "__main__":
    main()
